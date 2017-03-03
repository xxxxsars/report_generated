import os


os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'report.settings')
import django

django.setup()

from upload.models import Second_Data,New_Second_Data,Data
from openpyxl import load_workbook
import re
from django.db import connection


class Data_tans():
    def __init__(self,filename,content_year,content_month):
        self.filename = filename
        self.year = content_year
        self.month = content_month
        self.wb = load_workbook(filename="upload/" + self.filename)
        self.sheet = self.wb.get_sheet_names()
        self.ws = self.wb.get_sheet_by_name(self.sheet[0])


    def run(self):
        rows_num = (len(list(self.ws.rows)))
        col_num = len(list(self.ws.columns))

        Class_di  =  {"一":"1","二":"2","三":"3","四":"4","五":"5","六":"6","七":"7","八":"8","九":"9","十":"10","十一":"11","十二":"12","十三":"13","十四":"14" }
        for row in range(2, rows_num + 1):


                Class = (self.ws.cell(row=row, column=2).value)
                Name = (self.ws.cell(row=row,column=4).value)
                Num = (self.ws.cell(row=row,column=3).value)
                Absenteeism = (self.ws.cell(row=row,column=6).value)



                if type(Absenteeism) == int:
                    Class_first=re.search(r"(\w{1})年(\w{1,2})班",Class).group(1)

                    Class_second=re.search(r"(\w{1})年(\w{1,2})班",Class).group(2)
                    if len(Class_second)<2 and Class_second!="十":
                        Class = Class_di[Class_first]+"0"+Class_di[Class_second]
                    elif Class_second=="十":
                        Class = Class_di[Class_first] + Class_di[Class_second]
                    else:
                        Class = Class_di[Class_first]+Class_di[Class_second]




                    Second_Data.objects.create(Second_Class=Class,Second_name=Name,Second_num=Num,Second_absenteeism=Absenteeism)



    def trans_data(self):
        cursor = connection.cursor()
        cursor.execute("""SELECT
                            upload_second_data.Second_Class,
                            upload_second_data.Second_name,
                            upload_second_data.Second_num,
                            Sum(upload_second_data.Second_absenteeism) AS count_absenteeism
                            FROM upload_second_data
                            GROUP BY
                            upload_second_data.Second_name
                            ORDER BY
                            upload_second_data.Second_Class ASC""")
        for row in cursor.fetchall():
            if row[3]//2 >=1:
                penalty =row[3]//2
                handle = "警告%s次" % penalty
                content = "統計%s年%s月曠課達%s節" % (self.year, self.month, row[3])
                New_Second_Data.objects.create(Class=row[0],name=row[1],num=row[2],content=content,handle=handle)




    def tt(self):
        C_all = sorted(set(New_Second_Data.objects.values_list("N_Class", flat=True)))

        datas = []
        for c in C_all:
            d = New_Second_Data.objects.filter(Class=c)
            datas.append(d)
            print(d.name)


    def t2(self):
        C_all = sorted(set(Data.objects.values_list("Class", flat=True)))

        datas = []
        for c in C_all:
            d = Data.objects.filter(Class=c)
            datas.append(d)


if __name__ =="__main__":
    a = Data_tans("國一出勤統計 (5).xlsx","100","11")
    a.trans_data()
