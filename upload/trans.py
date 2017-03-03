import os


os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'report.settings')
import django

django.setup()





from openpyxl import load_workbook
import openpyxl
import re
from upload.models import Data




#匯入django 系統的模組 模組來源為root的settings



class Auto():
    def __init__(self,filename,content_year,content_month):
        self.filename = filename
        self.year = content_year
        self.month = content_month
        self.wb = load_workbook(filename="upload/"+self.filename)
        self.sheets = self.wb.get_sheet_names()
        self.ws = self.wb.get_sheet_by_name(self.sheets[0])
    def run(self):
        rows_num =len(list(self.ws.rows))
        cols_num =len(list(self.ws.columns))
        #班級range
        range_class = []
        #存入欄位名稱
        col_name = []
        #設定欄位從第五蘭開始
        start_col = 1




        #取出首行的欄位，存入col_name
        for col in range(start_col ,cols_num+1):
            col_name.append(self.ws.cell(row =1,column=col).value)
        # print(col_name)


        #取出同班級學生的資料
        for row in range(1,rows_num+1):
            value = (self.ws.cell(row=row,column=2).value)

            if re.search(r"\w{2}\d{1,2}",value):
                range_class.append(row)
        range_class.append(rows_num+1)
        #抓出的班級為類別的list 放入大list中
        all = []
        for i,v in enumerate(range_class):
            #先將資料撈成list便已和欄位結合為dict
            c_students= []
            #將每班的學生的全部dict轉存為list
            c_students2= []
            try:
            #c_student1操作
                for row in range(range_class[i]+1,range_class[i+1]):
                     for col in range(start_col ,cols_num):
                         c_students.append(self.ws.cell(row=row,column=col).value)
            except:pass

            if len(c_students) >0:
                all.append(c_students)



        #班級的雜亂資料整理 如中五1變為501
        tans_key = {"一":"1","二":"2","三":"3","四":"4","五":"5","六":"6","七":"7","八":"8","九":"9"}
        tans_class = []
        #移除最後一個元素
        range_class.pop()


        doc = open("test.csv",'w',encoding='utf-8')
        doc.write("班級"+','+"姓名" +','+"座號" +','+"獎勵事實" +','+"擬定處理"+"\n")
        for i in range_class:
            data = (self.ws.cell(row=i,column=2).value)
            if re.search(r"\w{2}(\d{1,2})", data):
                if len(re.search(r"\w{1}(\w{1})(\d{1,2})", data).group(2))>=2:
                    tans_class.append((tans_key[re.search(r"\w{1}(\w{1})",data).group(1)])+(re.search(r"\w{1}(\w{1})(\d{1,2})", data).group(2)))
                else:

                    tans_class.append((tans_key[re.search(r"\w{1}(\w{1})", data).group(1)]) +"0"+ (re.search(r"\w{1}(\w{1})(\d{1,2})", data).group(2)))


        for index,students in enumerate(all):
            c_students2 = []
            for i in range(len(students)):
                if i%(cols_num-1) ==0:
                    c_di = dict(zip(col_name,students[i:i+cols_num-1]))
                    c_di["班級"]=tans_class[index]
                    c_students2.append(c_di)

            for student in (c_students2):
                k_li =[]
                v_li =[]
                di_student={}



                for key,value in  (list(student.items())):
                    if (type(value)) ==int and value>0 and  key =="曠課" :

                        k_li.append(key)
                        v_li.append(value)
                        di_student.update({"姓名":student["姓名"],"班級":student["班級"],"座號":student["座號"]})
                tmp_di = dict(zip(k_li,v_li))
                di_student.update(tmp_di)
                if (di_student and di_student["曠課"]//2>=1):
                    penalty = (di_student["曠課"]//2)
                    handle = "警告%s次"%penalty
                    content = "統計%s年%s月曠課達%s節"%(self.year,self.month,di_student["曠課"])
                    doc.write(di_student["班級"]+','+di_student["姓名"]+','+di_student["座號"]+','+content+','+handle+"\n")

                    Data.objects.create(Class=di_student["班級"],name=di_student["姓名"],num=di_student["座號"],content=content,handle=handle)
                    print(content)


if __name__ =="__main__":
    a = Auto()
    a.run()


